"""
Build AppDev_Fischer_Tim/test-protocol.xlsx from docs/test-protocol.csv.

- Sheet "Test Cases" — alle Cases mit Spalten ID/UC/Title/Type/Framework/Preconditions/Steps/Test-Data/Expected/Priority
- Sheet "Summary"   — Inventar nach Type + pro UC + Coverage-Validierung

Run: python3 docs/_build_xlsx.py
"""
import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CSV_FILE = ROOT / "docs" / "test-protocol.csv"
XLSX_FILE = ROOT / "AppDev_Fischer_Tim" / "test-protocol.xlsx"

# UC-Zuordnung für Manual-Cases (xlsx-spezifisch, da CSV "Manual" als UC führt)
MANUAL_UC_MAP = {
    "M001": "UC06",  # Cage-Borders Chrome (Play-Screen)
    "M002": "UC06",  # Cage-Borders Firefox
    "M003": "UC06",  # Tastatur-Navigation
    "M004": "UC08",  # Highscore Cross-Browser
    "M005": "UC06",  # Mobile-Viewport (Game-Layout)
}

TYPE_LABEL = {
    "U": "Unit (xUnit)",
    "C": "Component (bUnit)",
    "I": "Integration (xUnit+WAF+MSSQL)",
    "E": "E2E (Playwright)",
    "M": "Manual",
}

PRIO_FILL = {
    "P1": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "P2": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "P3": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
}

HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(name="Inter", bold=True, color="FFFFFF")

# ----------------------------------------------------------------------------
# Read CSV
# ----------------------------------------------------------------------------
rows: list[dict[str, str]] = []
with CSV_FILE.open("r", encoding="utf-8-sig", newline="") as f:
    reader = csv.DictReader(f, delimiter=";")
    for r in reader:
        rows.append(r)

# Map Manual UC values from "Manual" → real UC reference
for r in rows:
    if r["UC"] == "Manual" and r["TestID"] in MANUAL_UC_MAP:
        r["UC"] = MANUAL_UC_MAP[r["TestID"]]
    if r["Type"] == "M":
        r["Framework"] = "Manual"

# Sort: by UC, then by numeric TestID
def sort_key(r):
    uc = r["UC"]
    tid = r["TestID"]
    num = int(re.sub(r"[^0-9]", "", tid) or 0)
    return (uc, tid[0], num)

rows.sort(key=sort_key)

# ----------------------------------------------------------------------------
# Workbook
# ----------------------------------------------------------------------------
wb = Workbook()

# === Sheet 1: Test Cases ===
ws = wb.active
ws.title = "Test Cases"

headers = ["ID", "UC", "Title", "Type", "Framework",
           "Preconditions", "Steps", "Test-Data", "Expected", "Priority"]
ws.append(headers)
for col_idx, _ in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")

for r in rows:
    ws.append([
        r["TestID"], r["UC"], r["Title"], r["Type"], r["Framework"],
        r["Preconditions"], r["Steps"], r["TestData"], r["Expected"], r["Priority"],
    ])

# Apply priority row-fill
for row_idx in range(2, ws.max_row + 1):
    prio = ws.cell(row=row_idx, column=10).value
    fill = PRIO_FILL.get(prio)
    if fill:
        for col_idx in range(1, 11):
            ws.cell(row=row_idx, column=col_idx).fill = fill

# Column widths
widths = [8, 7, 60, 6, 28, 32, 60, 26, 50, 9]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# Wrap text on multi-line columns
for row_idx in range(2, ws.max_row + 1):
    for col_idx in (3, 6, 7, 8, 9):
        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

# === Sheet 2: Summary ===
ws2 = wb.create_sheet("Summary")

# Type breakdown
type_counts = {"U": 0, "C": 0, "I": 0, "E": 0, "M": 0}
prio_counts = {"P1": 0, "P2": 0, "P3": 0}
uc_counts: dict[str, int] = {}
uc_per_kind: dict[str, dict[str, int]] = {}

for r in rows:
    t = r["Type"]
    type_counts[t] = type_counts.get(t, 0) + 1
    prio_counts[r["Priority"]] = prio_counts.get(r["Priority"], 0) + 1
    uc = r["UC"]
    uc_counts[uc] = uc_counts.get(uc, 0) + 1
    title = (r["Title"] or "").lower()
    kind = "pos"
    if title.startswith("negative") or title.startswith("neg") or "→ reject" in title or "neg:" in title:
        kind = "neg"
    elif title.startswith("boundary") or "boundary" in title:
        kind = "boundary"
    uc_per_kind.setdefault(uc, {"pos": 0, "neg": 0, "boundary": 0})
    uc_per_kind[uc][kind] = uc_per_kind[uc].get(kind, 0) + 1

total = len(rows)

def section_header(title: str):
    ws2.append([title, ""])
    cell = ws2.cell(row=ws2.max_row, column=1)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    ws2.cell(row=ws2.max_row, column=2).fill = HEADER_FILL

ws2.append(["Metric", "Value"])
ws2.cell(row=1, column=1).font = HEADER_FONT
ws2.cell(row=1, column=1).fill = HEADER_FILL
ws2.cell(row=1, column=2).font = HEADER_FONT
ws2.cell(row=1, column=2).fill = HEADER_FILL

ws2.append(["Total Test Cases", total])

section_header("Aufschlüsselung nach Type")
for k in ("U", "C", "I", "E", "M"):
    ws2.append([f"  {TYPE_LABEL[k]}", type_counts[k]])

section_header("Aufschlüsselung nach Priorität")
for k in ("P1", "P2", "P3"):
    ws2.append([f"  {k}", prio_counts[k]])

section_header("Tests pro Use Case")
for uc in sorted(uc_counts):
    ws2.append([f"  {uc}", uc_counts[uc]])

section_header("README §3.1 Coverage-Validierung (pos / neg / boundary)")
ws2.append(["UC", "pos / neg / boundary"])
for uc in sorted(uc_per_kind):
    k = uc_per_kind[uc]
    ws2.append([f"  {uc}", f"{k['pos']} / {k['neg']} / {k['boundary']}"])

ws2.column_dimensions["A"].width = 60
ws2.column_dimensions["B"].width = 28

# ----------------------------------------------------------------------------
# Save
# ----------------------------------------------------------------------------
XLSX_FILE.parent.mkdir(parents=True, exist_ok=True)
wb.save(XLSX_FILE)
print(f"Wrote {total} test cases to {XLSX_FILE}")
print(f"Type breakdown: {type_counts}")
print(f"Priority: {prio_counts}")
print(f"UCs: {sorted(uc_counts.keys())}")
